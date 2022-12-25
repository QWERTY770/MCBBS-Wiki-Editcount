import openpyxl as xl
from requests import get
from requests.exceptions import RequestException
from retrying import retry

import json
import os
from time import localtime
import threading


__author__ = "QWERTY_52_38"
__version__ = "0.5"
rev_api = "https://mcbbs.wiki/api.php?action=query&format=json&prop=revisions&revids="
# revisions api
folder = r"D:\python\mcbbswiki\MCBBS-Wiki-Editcount"  # this should be changed to your own directory
headers = {"user-agent": 'QWERTY770/1.0 Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36 Edg/107.0.1418.52'}
# avoid returning 418

namespace_score = {0: 3, 1: 0.125, 4: 1, 5: 0.125,
                   10: 2.5, 11: 0.125, 12: 2, 13: 0.125, 14: 1, 15: 0.125}

namespace_loca = {0:0, 1: 1, 4: 2, 5: 1, 10: 3, 
                   11: 1, 12: 4, 13: 1, 14: 5, 15: 1}  # main talk mcbbswiki template help category


@retry(stop_max_attempt_number=10)
def get_page(url: str):
    return get(url, timeout=5, headers=headers).text


def get_revs(start, end):
    print(f"{start} to {end} started!\n", end="")
    for i in range(start, end + 1):
        rev = get_page(rev_api + str(i))
        with open(os.path.join(os.path.join(folder, "rev"), f"rev_{i}.txt"), "w") as f:
            f.write(rev)
    print(f"{start} to {end} finished!\n", end="")


def get_edit_score_dic(start: int, end: int) -> dict:
    user_dic = {}
    for i in range(start, end + 1):
        with open(os.path.join(os.path.join(folder, "rev"), f"rev_{i}.txt"), "r") as f:
            js = json.loads(f.read())
            try:
                # print(js["query"]["pages"])
                page_id = list(js["query"]["pages"].keys())[0]
                namespace = js["query"]["pages"][page_id]["ns"]
                title = js["query"]["pages"][page_id]["title"]
                user = js["query"]["pages"][page_id]["revisions"][0]["user"]
            except:
                continue
            if user not in user_dic:
                user_dic[user] = [0,0,0,0,0,0,0,0] # main talk mcbbswiki template help category score total
            if namespace in namespace_score:
                user_dic[user][namespace_loca[namespace]] += 1
                user_dic[user][-2] += namespace_score[namespace]
            user_dic[user][-1] += 1
    return user_dic


def make_workbook(dic: dict, filename=f"mcbbswiki-useredit-{localtime().tm_year}{localtime().tm_mon}{localtime().tm_mday}-QWERTY770.xlsx"):
    wb = xl.Workbook()
    ws = wb.create_sheet('main',0)
    
    ws.cell(row=1, column=1).value = "用户名"
    ws.cell(row=1, column=2).value = "编辑总计"
    ws.cell(row=1, column=3).value = "（主）"
    ws.cell(row=1, column=4).value = "讨论"
    ws.cell(row=1, column=5).value = "MCBBS Wiki"
    ws.cell(row=1, column=6).value = "模板"
    ws.cell(row=1, column=7).value = "帮助"
    ws.cell(row=1, column=8).value = "分类"
    ws.cell(row=1, column=9).value = "编辑积分"

    for m, i in enumerate(dic.keys()):
        ws.cell(row=m+2, column=1).value = i
        ws.cell(row=m+2, column=2).value = dic[i][-1]
        for n, j in enumerate(dic[i]):
            ws.cell(row=m+2, column=n+3).value = j
    
    wb.save(os.path.join(folder, filename))
    wb.close()


if __name__ == "__main__":
    # get_revs(41025, 41029)
    make_workbook(get_edit_score_dic(1, 41029))
    print("Finished!")
